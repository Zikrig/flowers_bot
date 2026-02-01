from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import Dict, List
import os
from datetime import datetime


class OrderTemplate:
    def __init__(self, orders_dir: str = "orders"):
        self.orders_dir = orders_dir
        os.makedirs(orders_dir, exist_ok=True)
    
    def create_order_blank(self, order: Dict) -> str:
        """Создать бланк заказа"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"Заказ {order.get('order_number', '')}"
        
        # Заголовок
        ws.merge_cells('A1:D1')
        header_cell = ws['A1']
        header_cell.value = f"БЛАНК ЗАКАЗА №{order.get('order_number', '')}"
        header_cell.font = Font(size=16, bold=True)
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.fill = PatternFill(start_color="FFE6E6FA", end_color="FFE6E6FA", fill_type="solid")
        
        # Информация о заказе
        row = 3
        ws[f'A{row}'] = "Варианты букетов:"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        bouquets = order.get("bouquets", [])
        for bouquet in bouquets:
            variant_name = bouquet.get("variant_name", "")
            quantity = bouquet.get("quantity", 0)
            count = bouquet.get("count", 0)
            ws[f'A{row}'] = f"  • {variant_name} - {quantity} шт. - {count} букет"
            row += 1
        
        row += 1
        ws[f'A{row}'] = f"Дата самовывоза: {order.get('pickup_date', '')}"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = f"Время самовывоза: {order.get('pickup_time', '')}"
        ws[f'A{row}'].font = Font(bold=True)
        
        # Настройка ширины колонок
        ws.column_dimensions['A'].width = 50
        
        # Сохранение файла
        filename = f"order_{order.get('order_number', '')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.orders_dir, filename)
        wb.save(filepath)
        
        return filepath


